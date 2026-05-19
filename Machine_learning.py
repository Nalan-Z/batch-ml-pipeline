# Machine_learning.py
# Author: Nalan-Z
#
# Main pipeline: reads a batch Excel file, preprocesses data, runs ML classifiers
# (LDA + GaussianNB) across analytes, applies 4-parameter logistic regression (4PLR)
# curve fitting, computes deficiencies/borderlines, and writes a processed summary.
#
# Requires: config.yaml in the working directory (see config.yaml for path setup).
# Run: python Machine_learning.py   (file picker dialog will open)

import time
start_time = time.time()
import os, sys, csv, shutil, argparse
import scipy, numpy, matplotlib, pandas, sklearn
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import pandas as pd
from openpyxl import load_workbook
import yaml

# ---------------------------------------------------------------------------
# Load config
# ---------------------------------------------------------------------------
with open("config.yaml", "r") as _f:
    _cfg = yaml.safe_load(_f)

WORKING_DIR   = _cfg["paths"]["working_dir"]
RESULTS_DIR   = _cfg["paths"]["results_dir"]
LOG_DIR       = _cfg["paths"]["log_dir"]
HISTORICAL_CSV = _cfg["paths"]["historical_data"]
BATCH_TEMPLATE = _cfg["files"]["batch_template"]

# ---------------------------------------------------------------------------
# argparse
# ---------------------------------------------------------------------------
parser = argparse.ArgumentParser(
    prog='Machine_learning.py',
    formatter_class=argparse.RawDescriptionHelpFormatter,
    description='''\
Batch ML pipeline: preprocessing → LDA/GaussianNB prediction → 4PLR curve
fitting → deficiency/borderline classification → processed summary output.

Outputs:
  <batch>_processed_summary.xlsx  – summary with validation metrics and batch stats
  <batch>_filtered_results.xlsx – upload-ready P3 values
  <batch>_processed_uncertainty_plot.png – reference count vs uncertainty scatter
  Various CSV log entries written to LOG_DIR (see config.yaml)
''')

# ---------------------------------------------------------------------------
# Helper: append one row to a CSV log file
# ---------------------------------------------------------------------------
def log_csv(filepath, row):
    """Append a single row to a CSV log file. Creates the file if needed."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with open(filepath, 'a', newline='') as fh:
        csv.writer(fh).writerow(row)

# ---------------------------------------------------------------------------
# File selection
# ---------------------------------------------------------------------------
print("Starting...", time.time())
Tk().withdraw()
infile = askopenfilename()

base_name = os.path.basename(infile)
file_name_without_ext = os.path.splitext(base_name)[0]




# ---------------------------------------------------------------------------
# Extract data from workbook
# ---------------------------------------------------------------------------
print("Loading workbook...", time.time())
workbook = load_workbook(filename=infile, data_only=True)
print("Workbook loaded.", time.time())
sheet = workbook["Converter"]

data = []
parameter1 = 0 #use parameter

for row in sheet.iter_rows():
    for cell in row:
        if cell.value == "ref":
            ref_row, ref_col = cell.row, cell.column
            start_col = ref_col - parameter1
            for i in range(ref_row, sheet.max_row + 1):
                row_data = [sheet.cell(row=i, column=j).value
                            for j in range(start_col, ref_col + 1)]
                data.append(row_data)
                if sheet.cell(row=i, column=start_col).value == 0:
                    break

print("Data extracted.", time.time())

desired_cells_df = pd.DataFrame(data)
desired_cells_df = desired_cells_df.iloc[1:]  # drop header row
desired_cells_df = desired_cells_df.loc[
    desired_cells_df[desired_cells_df.columns[0]] != 0]

desired_cells_df.columns = ["Batch", "Age", "Gender", "Accession",...]
#add more if needed

output_filename = "extracted_" + file_name_without_ext + ".xlsx"
desired_cells_df.to_excel(output_filename, index=False)
infile = output_filename

print("Extraction done.", time.time())







# ---------------------------------------------------------------------------
# Preprocess
# ---------------------------------------------------------------------------
from scipy import stats

df = pandas.read_excel(infile)
base_name = os.path.basename(infile)
file_name_without_ext = os.path.splitext(base_name)[0]

print("Preprocess 1", time.time())

# Export cases with missing values
missing_value_cases = df[df.isnull().any(axis=1)]
missing_value_file = infile.replace('.xlsx', '_missing_values.xlsx')
missing_value_cases.to_excel(missing_value_file)

# Remove cases with extreme missingness
df = df.dropna(thresh=41)

# Write P1 file
outfile = infile.replace('.xlsx', 'file.xlsx')
df.to_excel(outfile)

print("Preprocess 2", time.time())
ref = df[['Age', 'cplumn', ...]]
ref.columns = ['Age_raw', 'column_raw',...]

# Fill missing numeric values with column means
numeric_cols = df.select_dtypes(include=['number']).columns
for col in numeric_cols:
    df.loc[:, col] = df[col].fillna(df[col].mean())

print("Preprocess 3", time.time())
# Z-score normalisation
# NOTE: Accession column must be non-numeric, otherwise drop it explicitly below.
numeric_cols = df.select_dtypes(include=[numpy.number]).columns
z_df = df[numeric_cols].apply(stats.zscore)
z_df = z_df.drop(['Batch', 'Accession'], axis=1)

# Replace zero z-scores with per-row mean (better imputation for missing analytes)
num = z_df.iloc[:, 1:44]
num[num.round(10) == 0] = numpy.nan
m = num.mean(axis=1)
for i, col in enumerate(num):
    num.iloc[:, i].fillna(m, inplace=True)
z_df = pandas.concat([z_df[['Age']], num], axis=1)

# Encode gender
metadata = df[['Batch', 'Accession', 'Gender']]
metadata['Gender'].replace(['M', 'F'], [1, 2], inplace=True)

print("Preprocess 3 – join processed data", time.time())
z_pre = pandas.concat([metadata, ref, z_df], axis=1)

plate1 = z_pre[['B1', 'B2', 'B3', 'B6', 'B12', ...]]
pl1_zscore = pandas.DataFrame({'PL1_zscore': plate1.mean(axis=1)})
z_predict = pandas.concat([z_pre, pl1_zscore], axis=1)

outfile = infile.replace('.xlsx', '_processed.csv')
z_predict.to_csv(outfile)











# ---------------------------------------------------------------------------
# ML imports
# ---------------------------------------------------------------------------
print("ML start", time.time())
from pandas.plotting import scatter_matrix
import matplotlib.pyplot as plt
from sklearn import model_selection
from sklearn.metrics import classification_report, confusion_matrix, accuracy_score
from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis, QuadraticDiscriminantAnalysis
from sklearn.naive_bayes import GaussianNB
from sklearn.svm import SVC
from sklearn.preprocessing import StandardScaler, PolynomialFeatures
from sklearn.pipeline import Pipeline
from numpy import array
from random import *
from scipy import stats



# ---------------------------------------------------------------------------
# Core ML function
# ---------------------------------------------------------------------------
def my_machine(i, learn, predict, df, predict_df,
               running_prediction, running_validation, running_Y_validation,
               estimator):
    analyte = df[learn]

    array_data = analyte.values
    X = array_data[:, :analyte.shape[1] - 1]
    Y = array_data[:, analyte.shape[1] - 1]
    validation_size = 0.15
    seed = int(randint(1, 10))
    X_train, X_validation, Y_train, Y_validation = model_selection.train_test_split(
        X, Y, test_size=validation_size, random_state=seed)

    estimator.fit(X_train, Y_train)
    validation_predictions = estimator.predict(X_validation)

    validation_predictions = pandas.DataFrame({i: validation_predictions})
    running_validation = pandas.concat([running_validation, validation_predictions],
                                       axis=1, sort=False)
    Y_validation = pandas.DataFrame({i: Y_validation})
    running_Y_validation = pandas.concat([running_Y_validation, Y_validation],
                                         axis=1, sort=False)

    predict_analyte = predict_df[predict]
    print(predict_analyte.describe())
    predict_array = predict_analyte.values
    predictions = pandas.DataFrame({i: estimator.predict(predict_array)})
    predict_analyte = predict_analyte.reset_index(drop=True)
    running_prediction = pandas.concat([running_prediction, predictions],
                                       axis=1, sort=False)
    return running_validation, running_prediction, running_Y_validation

# ---------------------------------------------------------------------------
# More preprocessing
# ---------------------------------------------------------------------------
infile = infile.replace('.xlsx', '_processed.csv')
learn_from = 85  # percent of historical cases used for training

from Run_files.predict_dict import *
from Run_files.learn_dict import *

df = pandas.read_csv(HISTORICAL_CSV)
print("Historical data loaded:", df.shape)

missing_value_cases = df[df.isnull().any(axis=1)]
df = df.dropna()
print("After dropping NaN:", df.shape)

learn_from = int(df.shape[0] * learn_from / 100)
df = df.sample(n=learn_from)
print("Training sample size:", df.shape)

predict_df = pandas.read_csv(infile)
print("Predict set shape:", predict_df.shape)

running_validation = pandas.DataFrame()
running_prediction = predict_df[['Batch', 'Accession', 'Age_raw',
                                  '100% Plate 1_raw', '2a_raw', '2b_raw']]
running_prediction = running_prediction.reset_index(drop=True)
running_Y_validation = pandas.DataFrame()

# ---------------------------------------------------------------------------
# Analyte value columns shared by both estimator loops
# ---------------------------------------------------------------------------
ANALYTE_COLS = [ "Vit A", "B1", "B12", "B2", "B3", "B6", "VitC", ...
]
META_COLS = ["Batch", "Accession", "Age_raw", "100% Plate 1_raw", "2a_raw", "2b_raw"]

def _run_estimator_loop(estimator, n_iter=10):
    """Run a single estimator n_iter times; return stacked predictions."""
    global running_validation, running_prediction, running_Y_validation
    loop_pred = []
    for _ in range(n_iter):
        for i in learn.keys():
            running_validation, running_prediction, running_Y_validation = my_machine(
                i, learn[i], predict[i], df, predict_df,
                running_prediction, running_validation, running_Y_validation, estimator)
        values = running_prediction[ANALYTE_COLS].values
        loop_pred.append(values)
        # reset for next iteration
        running_validation = pandas.DataFrame()
        running_prediction = predict_df[META_COLS].reset_index(drop=True)
        running_Y_validation = pandas.DataFrame()
    return loop_pred

print("ML – LinearDiscriminantAnalysis start", time.time())
loop_prediction = _run_estimator_loop(LinearDiscriminantAnalysis(), n_iter=10)

print("ML – GaussianNB start", time.time())
loop_prediction += _run_estimator_loop(GaussianNB(), n_iter=10)

lp_avg        = numpy.around(numpy.mean(loop_prediction, axis=0))
lp_variability = numpy.around(scipy.stats.iqr(loop_prediction, axis=0))

output_avg = pandas.DataFrame(lp_avg)





# ---------------------------------------------------------------------------
# Deficiency classification (bins → 1-5 labels; <3 = deficient)
# ---------------------------------------------------------------------------
ANALYTE_BINS = {
    # (analyte_col, bins, labels)
    'Zn':          ([20,36,38,42,44,65],   [1,2,3,4,5]),
    'B1':          ([52,77,79,83,85,126],  [1,2,3,4,5]),
    'B2':          ([35,52,54,58,60,94],   [1,2,3,4,5]),
    'B3':          ([52,79,81,85,87,132],  [1,2,3,4,5]),
    'B6':          ([30,53,55,59,61,96],   [1,2,3,4,5]),
    'B12':         ([0,14,15,18,20,40],    [1,2,3,4,5]), 
    #...
}

# Ordered list matches original column ordering for concat output
ANALYTE_ORDER = [
    'Vit A','B1','B12','B2','B3','B6','VitC', ...
]


def _classify_analyte(df, analyte, mode):
    """Return a boolean Series; mode='def' → <3, mode='bdl' → ==3."""
    bins, labels = ANALYTE_BINS[analyte]
    scored = pandas.to_numeric(
        pandas.cut(df[analyte], bins=bins, include_lowest=True, labels=labels))
    return scored < 3 if mode == 'def' else scored == 3


def _write_label_txt(df_bool, suffix, replace_str):
    """Write a label text file (Deficiencies.txt / Borderlines.txt style)."""
    fname = f"{suffix}.txt"
    col = pandas.DataFrame()
    analytes = df_bool.columns.tolist()
    for n, a in enumerate(analytes):
        col_name = f"{a}_{suffix.rstrip('s')}"  # e.g. Lipoic_def
        c = df_bool[a].replace([True, False], [col_name, 0])
        col = pandas.concat([col, c], axis=1)
    col.to_csv(fname, index=False, header=False)
    with open(fname, 'r') as fh:
        data = fh.read().replace('0', '').replace(',', '').replace(replace_str, ',').replace('CoQ1', 'CoQ10')
    with open(fname, 'w') as fh:
        fh.write(data)


def num_defs(df, outfile_name):
    """Classify deficiencies, write log and Deficiencies.txt, return count series."""
    def_flags = pandas.concat(
        {a: _classify_analyte(df, a, 'def') for a in ANALYTE_ORDER}, axis=1)

    # Log mean deficiency rates
    def_binary = def_flags.replace([True, False], [1, 0])
    log_csv(os.path.join(LOG_DIR, "Deficiency_log.csv"),
            [file_name_without_ext] + def_binary.mean().values.tolist())

    _write_label_txt(def_flags, 'Deficiencies', '_def')

    counts = (def_flags == True).sum(axis=1)
    return pandas.DataFrame({'Deficiencies': counts})


def num_BDLS(df, outfile_name):
    """Classify borderlines, write log and Borderlines.txt, return count series."""
    bdl_flags = pandas.concat(
        {a: _classify_analyte(df, a, 'bdl') for a in ANALYTE_ORDER}, axis=1)

    bdl_binary = bdl_flags.replace([True, False], [1, 0])
    log_csv(os.path.join(LOG_DIR, "Borderline_log.csv"),
            [file_name_without_ext] + bdl_binary.mean().values.tolist())

    _write_label_txt(bdl_flags, 'Borderlines', '_BDL')

    counts = (bdl_flags == True).sum(axis=1)
    return pandas.DataFrame({'Borderlines': counts})


def uncert(df):
    """Flag analytes with high variability and write uncert.txt."""
    v_cols = [f"{a}_V" for a in ANALYTE_ORDER]
    df2 = df[v_cols]
    threshold = df2.mean() + 2 * df2.std()
    tf = df2 > threshold
    col = pandas.DataFrame()
    for a in tf.columns:
        c = tf[a].replace([True, False], [a, 0])
        col = pandas.concat([col, c], axis=1)
    col.to_csv('uncert.txt', index=False, header=False)
    with open('uncert.txt', 'r') as fh:
        data = (fh.read()
                .replace('0,', '').replace(',0', '').replace('0', '').replace('_V', ''))
    with open('uncert.txt', 'w') as fh:
        fh.write(data)

# ---------------------------------------------------------------------------
# 4-Parameter Logistic Regression (4PLR)
# ---------------------------------------------------------------------------
from scipy.optimize import leastsq
from scipy.stats import zscore


def logistic4(x, A, B, C, D):
    """4PL logistic equation."""
    return D + ((A - D) / (1 + (abs(x / C) ** B)))


def residuals(p, y, x):
    A, B, C, D = p
    return y - logistic4(x, A, B, C, D)


def peval(x, p):
    A, B, C, D = p
    return logistic4(x, A, B, C, D)


def FourPLR(x, y_meas):
    p0 = [50, 4, 60, 50]
    plsq = leastsq(residuals, p0, args=(y_meas, x))
    return peval(x, plsq[0]), plsq[0]


# Per-analyte 4PLR configuration: (weight, z_col, log_filename)
FOURPLR_CONFIG = {
    'Vit A':       (0.13,  '2b_z', 'Vit_A.csv'),
    'B1':          (0.1,   '1_z',  'B1.csv'),
    'B12':         (0.15,  '1_z',  'B12.csv'),
    'B2':          (0.118, '1_z',  'B2.csv'),
    'B3':          (0.089, '1_z',  'B3.csv'),
    'B6':          (0.114, '1_z',  'B6.csv'),
    'VitC':        (0.127, '2a_z', 'VitC.csv'),
}

FOURPLR_ANALYTE_ORDER = ['Vit A','B1','B12','B2','B3','B6','VitC',...]


def FourPLR_Filter(final, base):
    cols_needed = (
        ["Batch", "Accession", "100% Plate 1_raw", "2a_raw", "2b_raw"] +
        [c for a in FOURPLR_ANALYTE_ORDER for c in (f"{a}_P1", a)]
        + ["CompoIndex1"]
    )
    df = final[cols_needed]

    mask = df.isnull().any(axis=1)
    df1 = df[~mask].copy()
    df2 = df[mask].copy()

    df1[['1_z', '2a_z', '2b_z']] = df1[
        ['100% Plate 1_raw', '2a_raw', '2b_raw']].apply(zscore)

    weighting = True
    adjust = 6  # reduce if weighting effect should be stronger

    if weighting:
        print('Analyte weights: ON  (adjust =', adjust, ')')
    else:
        print('Analyte weights: OFF')

    # --- first analyte initialises df1_sum; rest are concatenated ---
    first = True
    for analyte in FOURPLR_ANALYTE_ORDER:
        weight_val, z_col, log_fname = FOURPLR_CONFIG[analyte]
        x = df1[f"{analyte}_P1"]
        y_meas = df1[analyte]
        print(analyte)
        y_filt, p = FourPLR(x, y_meas)

        # log 4PLR curve parameters
        log_csv(os.path.join(LOG_DIR, log_fname),
                [final.iloc[0, 0], p[0], p[1], p[2], p[3]])

        if weighting:
            y_filt = y_filt * (weight_val / adjust) * df1[z_col] + y_filt

        col_df = pandas.DataFrame(y_filt, columns=[analyte])
        if first:
            df1_sum = col_df
            first = False
        else:
            df1_sum = pandas.concat([df1_sum, col_df], axis=1)

    df1_sum = df1_sum.round()

    df1_together = pandas.concat(
        [df1_sum, df1[["Batch", "Accession", "100% Plate 1_raw", "2a_raw", "2b_raw",
                        "CompoIndex1"] +
                       [f"{a}_P1" for a in FOURPLR_ANALYTE_ORDER] +
                       ["CompoIndex2_P1", "CompoIndex2"]],
        ], axis=1)

    ordered_cols = (
        ["Batch", "Accession", "100% Plate 1_raw", "2a_raw", "2b_raw"] +
        [c for a in FOURPLR_ANALYTE_ORDER for c in (f"{a}_P1", a)] +
        ["CompoIndex1"]
    )
    df1_sorted = df1_together[[c for c in ordered_cols if c in df1_together.columns]]

    df1_sorted = df1_sorted.assign(Low_GLC=85)  # complete case flag
    df2 = df2.assign(Low_GLC=84)                # incomplete case flag

    df1_and_df2 = df1_sorted._append(df2)
    df1_and_df2 = pandas.DataFrame.sort_index(df1_and_df2)

    df1_and_df2 = (df1_and_df2
                   .assign(Cysteine_1=1, Cysteine_3=1, CompoIndex2_1=95))

    P3_COLS = ["Batch", "Accession", "B1", "B2", "B3", "B6", "B12", "Folate",...]
    filt_P3 = df1_and_df2[P3_COLS].dropna()

    infile_no_ext = os.path.splitext(infile)[0]
    filt_P3.to_excel(infile_no_ext + '_filtered_results.xlsx', index=False)

    SUMMARY_COLS = [
        "Batch", "Accession", "Accession_P1", "Age_raw", "overall_validation",
        "100% Plate 1_raw", "2a_raw", "2b_raw"
    ] + [
        c for a in FOURPLR_ANALYTE_ORDER
        for c in (f"{a}_P1", a, f"{a}_V")
    ] + ["CompoIndex1", "CompoIndex1_V"]

    combined = pandas.concat(
        [df1_and_df2,
         final[["Accession_P1", "Age_raw", "overall_validation"] +
               [f"{a}_V" for a in FOURPLR_ANALYTE_ORDER] +
               ["CompoIndex1_V"]]], axis=1)

    combined_sorted = combined[[c for c in SUMMARY_COLS if c in combined.columns]]
    return combined_sorted

# ---------------------------------------------------------------------------
# Import intermediate results files
# ---------------------------------------------------------------------------
print("Importing intermediate results files...", time.time())

file_name = file_name_without_ext + "_processed_avg_results.xlsx"
results_df = pandas.read_excel(file_name)

file_name = file_name_without_ext + "_processed_variability_results.xlsx"
validation_df = pandas.read_excel(file_name)
print(validation_df.columns)

columns_to_drop = ["Batch", "Accession", "Age_raw",
                   "100% Plate 1_raw", "2a_raw", "2b_raw", 'Unnamed: 0']
validation_df = validation_df.drop(columns=columns_to_drop)

validation_df.columns = ["Vit A_V","B1_V","B12_V","B2_V","B3_V","B6_V", "VitC_V",...]
overall_validation = pandas.DataFrame(
    {'overall_validation': validation_df.mean(axis=1)})
validation_df = pandas.concat([validation_df, overall_validation], axis=1)

combined_results_df = pandas.concat([results_df, validation_df], axis=1)
combined_results_df = combined_results_df[[
    "Vit A","Vit A_V","B1","B1_V","B12","B12_V","B2","B2_V","B3","B3_V",
    "B6","B6_V","VitC","VitC_V",...]]

# Import P1 values
print("Importing P1 data...", time.time())
file_name = file_name_without_ext + "_P1.xlsx"
P1_df = pandas.read_excel(file_name)
P1_df = P1_df.drop(columns=['Unnamed: 0'])

P1_df.columns = [
    "Batch","Age_raw","Gender","B1_P1","B2_P1","B3_P1","B6_P1","B12_P1",...
]

P1_important_df = P1_df[[
    "Batch","Age_raw","Gender","B1_P1","B2_P1","B3_P1","B6_P1","B12_P1",...
]].reset_index(drop=True)

# ---------------------------------------------------------------------------
# Combine and run 4PLR filter
# ---------------------------------------------------------------------------
combined_values = pandas.concat([combined_results_df, P1_important_df], axis=1)

FINAL_COLS = [
    "Batch","Accession","Accession_P1","Age_raw",
    "Vit A_P1","Vit A","Vit A_V","B1_P1","B1","B1_V","B12_P1","B12","B12_V",
    "B2_P1","B2","B2_V","B3_P1","B3","B3_V","B6_P1","B6","B6_V",
    "VitC_P1","VitC","VitC_V","overall_validation"
]
final_combined_values = combined_values[FINAL_COLS]

# Handle missing-value cases
missing = infile.replace('.txt', '_missing_values.csv')
missing_df = pandas.read_csv(missing)
rows = missing_df[missing_df.isnull().sum(axis=1) < 8].index
missing_df.drop(rows, inplace=True)
care = missing_df[["Batch","Accession","Age","100% Plate 1","2a","2b"]]
care.columns = ["Batch","Accession","Age_raw","100% Plate 1_raw","2a_raw","2b_raw"]
care = care.reset_index(drop=True)
dummy_cols = [c for c in FINAL_COLS if c not in care.columns]
dummy_df = pandas.DataFrame(numpy.nan, index=numpy.arange(len(missing_df)),
                             columns=dummy_cols)
combined_missing_df = pandas.concat([care, dummy_df], axis=1)[FINAL_COLS]

everything = final_combined_values._append(combined_missing_df, sort=False)
final = everything.sort_values("Accession").reset_index(drop=True)
final = final[[c for c in FINAL_COLS if c != 'overall_validation'] + ['overall_validation']]  # keep ov at end for now

# 4PLR filter
combined_sorted = FourPLR_Filter(final, infile)
final = combined_sorted

# Deficiencies and borderlines
Deficiencies = num_defs(final, results_df)
defs = pandas.read_table('Deficiencies.txt', names=['Defs'], skip_blank_lines=False)
final = pandas.concat([final, Deficiencies, defs], axis=1)

Borderlines = num_BDLS(final, results_df)
BDLS = pandas.read_table('Borderlines.txt', names=['BDLS'], skip_blank_lines=False)
final = pandas.concat([final, Borderlines, BDLS], axis=1)

per_diff = (final['100% Plate 1_raw'] - final['2a_raw']) / final['100% Plate 1_raw'] * 100
final = pandas.concat([final, pandas.DataFrame({'Percent_Diff': per_diff})], axis=1)

uncert(final)
UNCERT = pandas.read_table('uncert.txt', names=['UNCERT'], skip_blank_lines=False)
final = pandas.concat([final, UNCERT], axis=1)

# ---------------------------------------------------------------------------
# Summary outputs
# ---------------------------------------------------------------------------
final_summary = final[[
    "Accession","Age_raw","overall_validation","Deficiencies","Defs",
    "CompoIndex1","CompoIndex1_V","CompoIndex2","CompoIndex2_V",
    "100% Plate 1_raw","2a_raw","2b_raw","Percent_Diff","BDLS"
]].copy()
final_summary['100% Plate 1_raw'] = final_summary['100% Plate 1_raw'].round()
final_summary['2a_raw'] = final_summary['2a_raw'].round()
final_summary['2b_raw'] = final_summary['2b_raw'].round()
final_summary.Percent_Diff = final_summary.Percent_Diff.round()
final_summary.overall_validation = final_summary.overall_validation.round(3)

analyte_summary = final[[
    "Accession",
    "Vit A_P1","Vit A","Vit A_V","B1_P1","B1","B1_V","B12_P1","B12","B12_V",
    "B2_P1","B2","B2_V","B3_P1","B3","B3_V","B6_P1","B6","B6_V",
    "Vit E_P1","Vit E","Vit E_V",
]]

for_batch_stats = final[[
    "P1_raw","2a_raw","2b_raw",
    "Vit A_P1","B1_P1","B12_P1","B2_P1","B3_P1","B6_P1","VitC_P1","Vit D3_P1",
    "Vit E_P1","K2_P1","Zn_P1"
]]
df_stats = for_batch_stats.describe().drop(['25%','50%','75%']).reindex(
    ['min','max','mean','std','count'])
trans = df_stats.T

# Log P1 means
mean_p1_log = [file_name_without_ext] + [
    final[c].mean() for c in [
        "100% Plate 1_raw","Vit A_P1","B1_P1","B12_P1","B2_P1","B3_P1","B6_P1","VitC_P1","Vit D3_P1",
        "Vit E_P1","K2_P1","Zn_P1"
    ]
]
log_csv(os.path.join(LOG_DIR, "Mean_p1_log.csv"), mean_p1_log)

outfile = infile.replace('.csv', '_processed_summary.xlsx')
with pandas.ExcelWriter(outfile) as writer:
    final_summary.to_excel(writer, sheet_name='Summary', index=False)
    analyte_summary.to_excel(writer, sheet_name='Analyte Results')
    trans.to_excel(writer, sheet_name='Batch Stats')
    final.to_excel(writer, sheet_name='Summary_for_AA')

# ---------------------------------------------------------------------------
# Uncertainty plot
# ---------------------------------------------------------------------------
plt.subplot(211)
plt.ylabel('Uncertainty Plate 1')
plt.title('Reference vs Uncertainty')
plt.scatter(final['100% Plate 1_raw'], final['overall_validation'])
plt.subplot(212)
plt.xlabel('100% Reference Count')
plt.ylabel('Uncertainty Plate 2')
plt.scatter(final['2a_raw'], final['overall_validation'])
plot_out = infile.replace('_processed.csv', '_processed_uncertainty_plot.png')
plt.savefig(os.path.join(WORKING_DIR, plot_out))

# Uncertainty log
ref = final["100% Plate 1_raw"]
below5000 = (ref < 5000).replace([True, False], [1, 0])
percent_below5000 = numpy.true_divide(below5000.sum(), below5000.count()) * 100

log_csv(os.path.join(LOG_DIR, "Batch_stats_log.csv"), [
    file_name_without_ext,
    final["Accession"].count(),
    final["overall_validation"].mean(),
    final["Age_raw"].mean(),
    final["Deficiencies"].mean(),
    final["100% Plate 1_raw"].mean(),
    final["2a_raw"].mean(),
    final["2b_raw"].mean(),
    final["100% Plate 1_raw"].max(),
    percent_below5000
])

# ---------------------------------------------------------------------------
# Precheck diagnostics
# ---------------------------------------------------------------------------
per_diff_desc = (final['100% Plate 1_raw'] - final['2a_raw']) / final['100% Plate 1_raw'] * 100
print('\n*******')
print(pandas.DataFrame({'Percent_Diff': per_diff_desc}).describe())
print('\n*******')
print(f"Average uncertainty: {final['overall_validation'].mean():.3f}  "
      f"({final['Accession'].count()} samples)")
print('*******\n')

print("Cases where 2b > 2a:")
print(final.Accession[final['2b_raw'] > final['2a_raw']].to_string(index=False))

print("\nHigh-error cases (|Percent_Diff| > 300):")
bad = final.Accession[abs(final['Percent_Diff']) > 300]
print("(none)" if bad.empty else bad.to_string(index=False))

# Swap detection
checkit = final[['Accession', '100% Plate 1_raw', '2a_raw']]
ch1 = checkit.iloc[1:].reset_index(drop=True)
ch2 = checkit.iloc[:-1].reset_index(drop=True)
merge = pandas.concat([ch1, ch2], axis=1)
merge.columns = ['Accession1','1a1','2a1','Accession2','1a2','2a2']
T1 = abs(merge['1a1']-merge['2a1'])/((merge['1a1']+merge['2a1'])/2)*100 >= 20
T2 = abs(merge['1a2']-merge['2a2'])/((merge['1a2']+merge['2a2'])/2)*100 >= 20
T3 = abs(merge['1a1']-merge['2a2'])/((merge['1a1']+merge['2a2'])/2)*100 <= 20
T4 = abs(merge['2a1']-merge['1a2'])/((merge['2a1']+merge['1a2'])/2)*100 <= 20
swaps = pandas.concat([T1,T2,T3,T4], axis=1).sum(axis=1) == 4
report = pandas.concat([pandas.DataFrame({'Swap': swaps}), merge], axis=1)
print("\nPossible plate swaps:")
print(report.Accession2[report.Swap].to_string(index=False),
      "↔",
      report.Accession1[report.Swap].to_string(index=False))

# ---------------------------------------------------------------------------
# Batch stats template fill
# ---------------------------------------------------------------------------
from shutil import copyfile
import openpyxl

processed_summary = pandas.read_excel(
    file_name_without_ext + "_processed_processed_summary.xlsx",
    sheet_name="Batch Stats")
selected_columns = processed_summary.iloc[:, 1:6]

new_file_path = 'batch_stats_copy.xlsx'
copyfile(BATCH_TEMPLATE, new_file_path)

workbook = openpyxl.load_workbook(new_file_path)
worksheet = workbook['Place Data Here']
for col_i, header in enumerate(selected_columns.columns.tolist()):
    worksheet.cell(row=1, column=col_i + 2, value=header)
for row_i, row_data in selected_columns.iterrows():
    for col_i, val in enumerate(row_data):
        worksheet.cell(row=row_i + 3, column=col_i + 2, value=val)
workbook.save(new_file_path)

# Conditional formatting on processed summary
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

ps_path = file_name_without_ext + "_processed_processed_summary.xlsx"
wb = load_workbook(ps_path)
ws = wb['Summary']
ws.freeze_panes = 'A2'
for col_idx, col in enumerate(ws.iter_cols(values_only=True, max_row=1)):
    if col[0] == "overall_validation":
        ov_col_letter = get_column_letter(col_idx + 1)
        break
ws.conditional_formatting.add(
    f'{ov_col_letter}2:{ov_col_letter}{ws.max_row}',
    ColorScaleRule(start_type='min', start_color='FF0000',
                   mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                   end_type='max', end_color='00FF00'))
wb.save(ps_path)

# ---------------------------------------------------------------------------
# Move intermediate files to results directory
# ---------------------------------------------------------------------------
os.makedirs(RESULTS_DIR, exist_ok=True)
FILE_EXTENSIONS = ["_P1.xlsx", "_missing_values.xlsx", "_avg_results.xlsx",
                   "_variability_results.xlsx", "_processed.csv",
                   "_filtered_results.xlsx"]
for fname in os.listdir(WORKING_DIR):
    if any(fname.endswith(ext) for ext in FILE_EXTENSIONS):
        shutil.move(os.path.join(WORKING_DIR, fname),
                    os.path.join(RESULTS_DIR, fname))

# Clean up temp files
for tmp in ["Deficiencies.txt", "Borderlines.txt", "uncert.txt"]:
    if os.path.exists(tmp):
        os.remove(tmp)

print('\n*******')
print('Done. Results written to:', RESULTS_DIR)
print(f"Runtime: {time.time() - start_time:.1f}s")
print('*******\n')
